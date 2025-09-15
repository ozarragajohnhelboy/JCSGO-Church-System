"""
Placeholder views for functions that need to be implemented
These are temporary implementations to keep the system working
while we organize the views systematically
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.contrib.auth import get_user_model
from django.utils import timezone
from datetime import timedelta

from ..models import *
from ..forms import *

User = get_user_model()

@login_required
def activity_logs(request):
    """View activity logs for the user's church"""
    user = request.user
    church = user.church
    
    action_filter = request.GET.get('action', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    activities = ActivityLog.objects.filter(
        church=church
    ).select_related('user', 'related_user').order_by('-timestamp')
    
    if action_filter:
        activities = activities.filter(action=action_filter)
    
    if user_filter:
        activities = activities.filter(user__id=user_filter)
    
    if date_from:
        activities = activities.filter(timestamp__date__gte=date_from)
    
    if date_to:
        activities = activities.filter(timestamp__date__lte=date_to)
    
    paginator = Paginator(activities, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    action_choices = [
        ('LOGIN', 'Login'),
        ('LOGOUT', 'Logout'),
        ('NEW_FRIEND_ADDED', 'New Friend Added'),
        ('REGULAR_MEMBER_ADDED', 'Regular Member Added'),
        ('STATUS_CHANGE', 'Status Change'),
        ('ROLE_UPDATED', 'Role Updated'),
        ('GROUP_JOINED', 'Group Joined'),
        ('GROUP_LEFT', 'Group Left'),
        ('ATTENDANCE_RECORDED', 'Attendance Recorded'),
    ]
    
    church_users = CustomUser.objects.filter(church=church, is_active=True).order_by('first_name')
    
    context = {
        'page_obj': page_obj,
        'action_filter': action_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
        'action_choices': action_choices,
        'church_users': church_users,
        'total_activities': activities.count(),
    }
    
    return render(request, 'members/logs/activity_logs.html', context)

@login_required  
def church_statistics(request):
    """Church statistics dashboard"""
    return render(request, 'members/reports/church_statistics.html', {})

@csrf_exempt
@login_required
def ajax_get_available_members(request, group_id):
    """AJAX endpoint to get available members for a group"""
    try:
        group = get_object_or_404(Group, pk=group_id)
        
        # Check if user has permission to add members to this group
        if group.leader != request.user and request.user.role.name != 'ADMIN':
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        # Get available members (regular members not in any group)
        available_members = CustomUser.objects.filter(
            church=group.church,
            is_active=True,
            is_new_friend=False,
            role__name__in=['VSL', 'CSL', 'CL', 'CM']
        ).exclude(
            regular_member_profile__group__isnull=False
        ).order_by('first_name', 'last_name')
        
        members_data = []
        for member in available_members:
            members_data.append({
                'id': member.id,
                'name': member.full_name,
                'email': member.email,
                'role': member.role.name if member.role else 'No Role'
            })
        
        return JsonResponse({'members': members_data})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
def ajax_update_timer_status(request, user_id):
    """AJAX endpoint to update timer status"""
    return JsonResponse({'success': True})

@csrf_exempt
@login_required
def ajax_record_attendance(request, user_id):
    """AJAX endpoint to record attendance"""
    return JsonResponse({'success': True})

@csrf_exempt
@login_required
def ajax_update_follow_up(request, new_friend_id):
    """AJAX endpoint to update follow up status"""
    if request.method == 'POST':
        try:
            new_friend = get_object_or_404(NewFriend, id=new_friend_id)
            
            if not request.user.can_access_church_data(new_friend.user.church):
                return JsonResponse({'error': 'Permission denied'}, status=403)
            
            status = request.POST.get('status')
            notes = request.POST.get('notes', '')
            
            new_friend.update_follow_up(status, notes)
            
            return JsonResponse({
                'success': True,
                'follow_up_status': new_friend.follow_up_status,
                'last_follow_up': new_friend.last_follow_up.isoformat() if new_friend.last_follow_up else None
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@csrf_exempt
@login_required
def ajax_add_to_group(request, user_id, group_id):
    """AJAX endpoint to add member to group"""
    return JsonResponse({'success': True})

@csrf_exempt
@login_required
def ajax_remove_from_group(request, user_id, group_id):
    """AJAX endpoint to remove member from group"""
    return JsonResponse({'success': True})

@login_required
def ajax_activity_details(request, activity_id):
    """AJAX endpoint to get activity details"""
    try:
        activity = get_object_or_404(ActivityLog, pk=activity_id)
        
        # Check if user has permission to view this activity
        if not request.user.can_access_church_data(activity.church):
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        # Prepare the details data
        details_html = f"""
        <div class="row">
            <div class="col-md-6">
                <h6 class="text-muted mb-3">Activity Information</h6>
                <table class="table table-sm">
                    <tr>
                        <td><strong>Action:</strong></td>
                        <td>
                            <span class="badge bg-{'success' if activity.action == 'LOGIN' else 'secondary' if activity.action == 'LOGOUT' else 'primary' if activity.action == 'REGISTER' else 'info' if activity.action == 'ATTENDANCE' else 'warning'}">
                                {activity.get_action_display()}
                            </span>
                        </td>
                    </tr>
                    <tr>
                        <td><strong>User:</strong></td>
                        <td>
                            <div class="d-flex align-items-center">
                                {'<img src="' + activity.user.profile_picture.url + '" alt="Profile" class="rounded-circle me-2" width="24" height="24">' if activity.user.profile_picture else '<div class="bg-secondary rounded-circle me-2 d-flex align-items-center justify-content-center" style="width: 24px; height: 24px;"><i class="bi bi-person text-white" style="font-size: 12px;"></i></div>'}
                                <span>{activity.user.full_name}</span>
                            </div>
                        </td>
                    </tr>
                    <tr>
                        <td><strong>Church:</strong></td>
                        <td>{activity.church.name if activity.church else 'N/A'}</td>
                    </tr>
                    <tr>
                        <td><strong>Timestamp:</strong></td>
                        <td>
                            {activity.timestamp.strftime('%B %d, %Y at %I:%M %p')}
                            <br><small class="text-muted">{activity.timestamp.strftime('%A')}</small>
                        </td>
                    </tr>
                </table>
            </div>
            <div class="col-md-6">
                <h6 class="text-muted mb-3">Technical Details</h6>
                <table class="table table-sm">
                    <tr>
                        <td><strong>IP Address:</strong></td>
                        <td>
                            <code>{activity.ip_address if activity.ip_address else 'N/A'}</code>
                        </td>
                    </tr>
                    <tr>
                        <td><strong>User Agent:</strong></td>
                        <td>
                            <small class="text-muted">
                                {activity.user_agent[:100] + '...' if activity.user_agent and len(activity.user_agent) > 100 else activity.user_agent or 'N/A'}
                            </small>
                        </td>
                    </tr>
                    {'<tr><td><strong>Related User:</strong></td><td>' + activity.related_user.full_name + '</td></tr>' if activity.related_user else ''}
                </table>
            </div>
        </div>
        
        <div class="row mt-3">
            <div class="col-12">
                <h6 class="text-muted mb-3">Description</h6>
                <div class="alert alert-light">
                    <p class="mb-0">{activity.description}</p>
                </div>
            </div>
        </div>
        
        {'<div class="row mt-3"><div class="col-12"><h6 class="text-muted mb-3">Additional Data</h6><div class="alert alert-info"><pre class="mb-0">' + str(activity.metadata) + '</pre></div></div></div>' if activity.metadata else ''}
        """
        
        return JsonResponse({
            'success': True,
            'html': details_html
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def export_members(request):
    """Export members data"""
    user = request.user
    church = user.church
    
    if not user.is_staff and not user.role.name in ['SUPER_ADMIN', 'ADMIN']:
        messages.error(request, 'You do not have permission to export data.')
        return redirect('members:member_list')
    
    export_format = request.GET.get('format', 'csv')
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    
    # Start with base queryset
    members = CustomUser.objects.filter(church=church, is_active=True)
    
    # Apply status filter
    if status == 'new_friends':
        members = members.filter(is_new_friend=True)
        filename_prefix = "new_friends"
    elif status == 'regular_members':
        members = members.filter(is_new_friend=False)
        filename_prefix = "regular_members"
    else:
        filename_prefix = "members"
    
    # Apply search filter (same logic as role_management view)
    if search:
        members = members.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Apply role filter
    if role_filter:
        members = members.filter(role__name=role_filter)
    
    # Order results
    members = members.order_by('role__name', 'first_name', 'last_name')
    
    from django.http import HttpResponse
    import csv
    
    # Prepare data for export
    data = []
    for member in members:
        data.append({
            'First Name': member.first_name,
            'Last Name': member.last_name,
            'Email': member.email,
            'Phone': member.phone_number or '',
            'Role': member.role.get_name_display() if member.role else '',
            'Member Type': 'New Friend' if member.is_new_friend else 'Regular Member',
            'Date Joined': member.date_joined.strftime('%Y-%m-%d'),
            'Last Attendance': member.last_attendance.strftime('%Y-%m-%d %H:%M') if member.last_attendance else ''
        })
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{church.domain}_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        headers = ['First Name', 'Last Name', 'Email', 'Phone', 'Role', 'Member Type', 'Date Joined', 'Last Attendance']
        writer.writerow(headers)
        
        for row_data in data:
            row = [row_data[header] for header in headers]
            writer.writerow(row)
        
        return response
    
    elif export_format == 'xlsx':
        import pandas as pd
        from io import BytesIO
        
        # Create DataFrame from data
        df = pd.DataFrame(data)
        
        # Create in-memory output file for Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Members')
            
            # Get the workbook and worksheet to apply formatting
            workbook = writer.book
            worksheet = writer.sheets['Members']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Set response content
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{church.domain}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    else:
        # Default to CSV if other format requested
        return redirect(f'{request.path}?format=csv&status={status}&search={search}&role={role_filter}')
    
    return response

@login_required
def export_role_data(request):
    """Export role management specific data"""
    user = request.user
    church = user.church
    
    if not user.is_staff and not user.role.name in ['SUPER_ADMIN', 'ADMIN']:
        messages.error(request, 'You do not have permission to export data.')
        return redirect('members:role_management')
    
    search = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    
    # Get role statistics
    roles = Role.objects.all()
    role_stats = {}
    
    for role in roles:
        count = CustomUser.objects.filter(church=church, role=role, is_active=True).count()
        total_users = CustomUser.objects.filter(church=church, is_active=True).count()
        percentage = round((count / total_users * 100) if total_users > 0 else 0, 1)
        
        role_stats[role.name] = {
            'name': role.get_name_display(),
            'count': count,
            'percentage': percentage
        }
    
    # Get users with role information
    users = CustomUser.objects.filter(church=church, is_active=True).select_related('role')
    
    # Apply filters (same as role_management view)
    if search:
        users = users.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search)
        )
    
    if role_filter:
        users = users.filter(role__name=role_filter)
    
    users = users.order_by('role__name', 'first_name', 'last_name')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="role_management_{church.domain}_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    # Write role statistics first
    writer.writerow(['ROLE STATISTICS'])
    writer.writerow(['Role', 'Count', 'Percentage'])
    for role_name, stats in role_stats.items():
        writer.writerow([stats['name'], stats['count'], f"{stats['percentage']}%"])
    
    writer.writerow([])  # Empty row separator
    
    # Write user role details
    writer.writerow(['USER ROLE DETAILS'])
    headers = ['Full Name', 'Email', 'Phone', 'Current Role', 'Member Status', 'Date Joined', 'Last Activity', 'Timer Status']
    writer.writerow(headers)
    
    for user_item in users:
        # Get last attendance
        last_attendance = Attendance.objects.filter(user=user_item).order_by('-time_in').first()
        last_activity = last_attendance.time_in.strftime('%Y-%m-%d %H:%M') if last_attendance else 'Never'
        
        # Timer status for new friends
        timer_status = ''
        if user_item.is_new_friend and hasattr(user_item, 'timer_status'):
            timer_status = f"{user_item.timer_status}{'st' if user_item.timer_status == 1 else 'nd' if user_item.timer_status == 2 else 'rd' if user_item.timer_status == 3 else 'th'} Timer"
        
        row = [
            user_item.get_full_name(),
            user_item.email,
            user_item.phone_number or '',
            user_item.role.get_name_display() if user_item.role else 'No Role',
            'New Friend' if user_item.is_new_friend else 'Regular Member',
            user_item.date_joined.strftime('%Y-%m-%d'),
            last_activity,
            timer_status
        ]
        writer.writerow(row)
    
    return response

@login_required
def role_management(request):
    """Role management view for admins only"""
    user = request.user

    if not (user.is_superuser or user.role.name == 'ADMIN'):
        messages.error(request, 'You do not have permission to access role management.')
        return redirect('churches:dashboard')
    
    church = user.church
    
    search = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    
    users = CustomUser.objects.filter(church=church, is_active=True)
    
    if search:
        users = users.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search)
        )
    
    if role_filter:
        users = users.filter(role__name=role_filter)
    
    users = users.order_by('role__name', 'first_name', 'last_name')
    
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    roles = Role.objects.filter(
        name__in=['VSL', 'CSL', 'CL', 'CM', 'NEW_FRIEND']
    ).order_by('name')
    
    role_stats = {}
    for role in roles:
        count = users.filter(role=role).count()
        if count > 0:
            role_stats[role.name] = {
                'name': role.get_name_display(),
                'count': count,
                'percentage': round((count / users.count() * 100) if users.count() > 0 else 0, 1)
            }
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'role_filter': role_filter,
        'roles': roles,
        'role_stats': role_stats,
        'total_users': users.count(),
    }
    
    return render(request, 'members/members/role_management.html', context)

@login_required
def ajax_update_user_role(request, user_id):
    """AJAX endpoint to update user role"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    user = request.user
    
    if not (user.is_superuser or user.role.name == 'ADMIN'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        target_user = get_object_or_404(CustomUser, pk=user_id, church=user.church)
        new_role_name = request.POST.get('role')
        
        if not new_role_name:
            return JsonResponse({'error': 'Role is required'}, status=400)
        
        new_role = get_object_or_404(Role, name=new_role_name)

        old_role = target_user.role
        target_user.role = new_role
        target_user.save()
        
        ActivityLog.objects.create(
            user=user,
            action='ROLE_CHANGE',
            description=f'Changed role from {old_role.get_name_display() if old_role else "None"} to {new_role.get_name_display()} for {target_user.full_name}',
            related_user=target_user,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        if target_user.is_new_friend and new_role_name not in ['NEW_FRIEND']:
            target_user.is_new_friend = False
            target_user.transition_date = timezone.now()
            target_user.save()
            
            RegularMember.objects.get_or_create(
                user=target_user,
                defaults={'role_type': new_role_name}
            )

            try:
                if hasattr(target_user, 'new_friend_profile'):
                    target_user.new_friend_profile.delete()
            except NewFriend.DoesNotExist:
                pass

        return JsonResponse({
            'success': True,
            'message': f'Role updated to {new_role.get_name_display()}',
            'new_role': new_role.get_name_display()
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def ajax_get_user_details(request, user_id):
    """AJAX endpoint to get user details for role management"""
    user = request.user
    
    if not (user.is_superuser or user.role.name == 'ADMIN'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        target_user = get_object_or_404(CustomUser, pk=user_id, church=user.church)
        
        group_info = None
        if not target_user.is_new_friend:
            try:
                regular_profile = RegularMember.objects.get(user=target_user)
                if regular_profile.group:
                    group_info = {
                        'id': regular_profile.group.id,
                        'name': regular_profile.group.name,
                        'type': regular_profile.group.get_group_type_display()
                    }
            except RegularMember.DoesNotExist:
                pass
        
        recent_activity = target_user.activity_logs.order_by('-timestamp')[:5]
        activity_list = []
        for activity in recent_activity:
            activity_list.append({
                'action': activity.get_action_display(),
                'description': activity.description,
                'timestamp': activity.timestamp.strftime('%Y-%m-%d %H:%M')
            })
        
        return JsonResponse({
            'success': True,
            'user': {
                'id': target_user.id,
                'full_name': target_user.full_name,
                'email': target_user.email,
                'phone_number': target_user.phone_number or 'Not provided',
                'current_role': target_user.role.get_name_display() if target_user.role else 'No role assigned',
                'is_new_friend': target_user.is_new_friend,
                'timer_status': target_user.timer_status if target_user.is_new_friend else None,
                'date_joined': target_user.date_joined.strftime('%Y-%m-%d'),
                'last_attendance': target_user.last_attendance.strftime('%Y-%m-%d %H:%M') if target_user.last_attendance else 'Never',
                'group': group_info,
                'recent_activity': activity_list
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def ajax_bulk_role_update(request):
    """AJAX endpoint for bulk role updates"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    user = request.user
    
    if not (user.is_superuser or user.role.name == 'ADMIN'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        user_ids = request.POST.getlist('user_ids[]')
        new_role_name = request.POST.get('role')
        
        if not user_ids or not new_role_name:
            return JsonResponse({'error': 'User IDs and role are required'}, status=400)
        
        new_role = get_object_or_404(Role, name=new_role_name)
        
        updated_count = 0
        for user_id in user_ids:
            try:
                target_user = CustomUser.objects.get(pk=user_id, church=user.church)
                old_role = target_user.role
                target_user.role = new_role
                target_user.save()
                
                ActivityLog.objects.create(
                    user=user,
                    action='ROLE_CHANGE',
                    description=f'Bulk update: Changed role from {old_role.get_name_display() if old_role else "None"} to {new_role.get_name_display()} for {target_user.full_name}',
                    related_user=target_user,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                updated_count += 1
                
            except CustomUser.DoesNotExist:
                continue
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully updated {updated_count} users to {new_role.get_name_display()}',
            'updated_count': updated_count
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def care_group_list(request):
    """List care groups for leadership roles (VSL, CSL, CL) and admin"""
    user = request.user

    if user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to access care groups.')
        return redirect('churches:dashboard')
    
    church = user.church

    search = request.GET.get('search', '')

    care_groups = Group.objects.filter(
        church=church,
        group_type='CARE',
        is_active=True
    ).select_related('leader').prefetch_related('members')

    if user.role.name == 'VSL':
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        care_groups = care_groups.filter(
            Q(leader=user) | 
            Q(leader__in=member_users)
        ).distinct()
    elif user.role.name == 'CSL':
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        care_groups = care_groups.filter(
            Q(leader=user) | 
            Q(leader__in=member_users)
        ).distinct()
    elif user.role.name == 'CL':
        care_groups = care_groups.filter(leader=user)
    elif user.role.name == 'ADMIN':
        pass
    
    if search:
        care_groups = care_groups.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(leader__first_name__icontains=search) |
            Q(leader__last_name__icontains=search)
        )

    care_groups = care_groups.order_by('name')

    paginator = Paginator(care_groups, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    led_groups = Group.objects.filter(
        church=church,
        group_type='CARE',
        leader=user,
        is_active=True
    )
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'total_care_groups': care_groups.count(),
        'led_groups_count': led_groups.count(),
        'can_create_group': user.role.name in ['VSL', 'CSL', 'CL', 'ADMIN'],
        'user_role': user.role.name,
    }
    
    return render(request, 'members/groups/care_group_list.html', context)

@login_required
def care_group_create(request):
    """Create a new care group - for VSL, CSL, CL roles and admin"""
    user = request.user

    if user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to create care groups.')
        return redirect('members:care_group_list')
    
    if request.method == 'POST':
        form = CareGroupForm(request.POST, church=user.church, user=user)
        if form.is_valid():
            care_group = form.save()

            ActivityLog.objects.create(
                user=user,
                church=user.church,
                action='GROUP_CREATED',
                description=f'Created care group: {care_group.name}'
            )
            
            messages.success(request, f'Care group "{care_group.name}" created successfully!')
            return redirect('members:care_group_detail', group_id=care_group.pk)
    else:
        form = CareGroupForm(church=user.church, user=user)
    
    context = {
        'form': form,
        'user_role': user.role.name,
    }
    
    return render(request, 'members/groups/care_group_create.html', context)

@login_required
def care_group_detail(request, group_id):
    """Detailed view of a care group"""
    user = request.user
    care_group = get_object_or_404(Group, pk=group_id, group_type='CARE')
    
    if not user.can_access_church_data(care_group.church):
        messages.error(request, 'You do not have permission to view this care group.')
        return redirect('members:care_group_list')
    
    if user.role.name in ['CSL', 'CL'] and care_group.leader != user:
        try:
            user_member_profile = RegularMember.objects.get(user=user)
            if user_member_profile.group != care_group:
                messages.error(request, 'You can only view care groups you lead or are a member of.')
                return redirect('members:care_group_list')
        except RegularMember.DoesNotExist:
            messages.error(request, 'You can only view care groups you lead or are a member of.')
            return redirect('members:care_group_list')

    members = care_group.members.select_related('user').order_by('user__first_name')

    recent_activity = ActivityLog.objects.filter(
        user__regular_member_profile__group=care_group
    ).select_related('user').order_by('-timestamp')[:10]

    available_members = None
    if care_group.leader == user or user.role.name == 'ADMIN':
        available_members = CustomUser.objects.filter(
            church=care_group.church,
            is_active=True,
            is_new_friend=False,
            role__name__in=['VSL', 'CSL', 'CL', 'CM']
        ).exclude(
            regular_member_profile__group__isnull=False
        ).order_by('first_name', 'last_name')
    
    context = {
        'group': care_group,
        'members': members,
        'recent_activity': recent_activity,
        'capacity_percentage': care_group.capacity_percentage,
        'is_full': care_group.is_full,
        'is_leader': care_group.leader == user or user.role.name == 'ADMIN',
        'available_members': available_members,
        'user_role': user.role.name,
    }
    
    return render(request, 'members/groups/care_group_detail.html', context)

@login_required
def care_group_edit(request, group_id):
    """Edit a care group - only the leader or admin can edit"""
    user = request.user
    care_group = get_object_or_404(Group, pk=group_id, group_type='CARE')

    if care_group.leader != user and user.role.name != 'ADMIN':
        messages.error(request, 'Only the group leader or admin can edit this care group.')
        return redirect('members:care_group_detail', group_id=group_id)
    
    if request.method == 'POST':
        form = CareGroupForm(request.POST, instance=care_group, church=user.church, user=user)
        if form.is_valid():
            updated_group = form.save()
 
            ActivityLog.objects.create(
                user=user,
                church=user.church,
                action='GROUP_UPDATED',
                description=f'Updated care group: {updated_group.name}'
            )
            
            messages.success(request, f'Care group "{updated_group.name}" updated successfully!')
            return redirect('members:care_group_detail', group_id=group_id)
    else:
        form = CareGroupForm(instance=care_group, church=user.church, user=user)
    
    context = {
        'form': form,
        'group': care_group,
        'user_role': user.role.name,
    }
    
    return render(request, 'members/groups/care_group_edit.html', context)

@login_required
def care_group_add_member(request, group_id):
    """Add a member to a care group"""
    user = request.user
    care_group = get_object_or_404(Group, pk=group_id, group_type='CARE')

    if care_group.leader != user and user.role.name != 'ADMIN':
        messages.error(request, 'Only the group leader or admin can add members to this care group.')
        return redirect('members:care_group_detail', group_id=group_id)
    
    if care_group.is_full:
        messages.error(request, 'This care group is already at full capacity.')
        return redirect('members:care_group_detail', group_id=group_id)
    
    if request.method == 'POST':
        member_id = request.POST.get('member_id')
        try:
            member_user = CustomUser.objects.get(pk=member_id, church=care_group.church)

            regular_member, created = RegularMember.objects.get_or_create(
                user=member_user,
                defaults={'role_type': member_user.role.name}
            )

            if regular_member.group:
                messages.error(request, f'{member_user.full_name} is already in a care group.')
                return redirect('members:care_group_detail', group_id=group_id)

            regular_member.group = care_group
            regular_member.save()

            ActivityLog.objects.create(
                user=user,
                church=user.church,
                action='GROUP_JOINED',
                description=f'{member_user.full_name} joined care group: {care_group.name}',
                related_user=member_user
            )

            messages.success(request, f'{member_user.full_name} has been added to {care_group.name}.')
            
        except CustomUser.DoesNotExist:
            messages.error(request, 'Selected member not found.')
        except Exception as e:
            messages.error(request, f'Error adding member: {str(e)}')
    
    return redirect('members:care_group_detail', group_id=group_id)

@login_required
def care_group_remove_member(request, group_id, member_id):
    """Remove a member from a care group"""
    user = request.user
    care_group = get_object_or_404(Group, pk=group_id, group_type='CARE')

    if care_group.leader != user and user.role.name != 'ADMIN':
        messages.error(request, 'Only the group leader or admin can remove members from this care group.')
        return redirect('members:care_group_detail', group_id=group_id)
    
    try:
        # member_id is actually the user's ID, not the RegularMember's ID
        member_user = get_object_or_404(CustomUser, pk=member_id, church=care_group.church)
        regular_member = get_object_or_404(RegularMember, user=member_user, group=care_group)

        regular_member.group = None
        regular_member.save()

        ActivityLog.objects.create(
            user=user,
            church=user.church,
            action='GROUP_LEFT',
            description=f'{member_user.full_name} left care group: {care_group.name}',
            related_user=member_user
        )

        messages.success(request, f'{member_user.full_name} has been removed from {care_group.name}.')
        
    except Exception as e:
        messages.error(request, f'Error removing member: {str(e)}')
    
    return redirect('members:care_group_detail', group_id=group_id)

@login_required
def role_new_friends_list(request):
    """List new friends endorsed to the current user (VSL, CSL, CL, CM)"""
    user = request.user

    if user.role.name not in ['VSL', 'CSL', 'CL', 'CM']:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('churches:dashboard')
    
    church = user.church

    search = request.GET.get('search', '')
    follow_up_status = request.GET.get('follow_up_status', '')
    timer_status = request.GET.get('timer_status', '')

    new_friends_users = CustomUser.objects.filter(
        church=church,
        is_active=True,
        is_new_friend=True,
        new_friend_profile__endorsed_to=user  
    )

    if search:
        new_friends_users = new_friends_users.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search) |
            Q(new_friend_profile__invited_by__first_name__icontains=search) |
            Q(new_friend_profile__invited_by__last_name__icontains=search)
        )
    
    if timer_status:
        new_friends_users = new_friends_users.filter(timer_status=timer_status)

    new_friends = []
    for user_obj in new_friends_users:
        try:
            new_friend_profile = NewFriend.objects.get(user=user_obj)
            if follow_up_status and new_friend_profile.follow_up_status != follow_up_status:
                continue
            new_friends.append(new_friend_profile)
        except NewFriend.DoesNotExist:
            new_friend_profile = NewFriend.objects.create(
                user=user_obj,
                invited_by=None,
                endorsed_to=user,
                notes='',
                is_active=True
            )
            new_friends.append(new_friend_profile)

    if follow_up_status:
        new_friends = [nf for nf in new_friends if nf.follow_up_status == follow_up_status]

    new_friends.sort(key=lambda x: x.registration_date, reverse=True)

    paginator = Paginator(new_friends, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'follow_up_status': follow_up_status,
        'timer_status': timer_status,
        'total_new_friends': len(new_friends),
        'pending_follow_up': len([nf for nf in new_friends if nf.follow_up_status == 'PENDING']),
        'engaged_count': len([nf for nf in new_friends if nf.follow_up_status == 'ENGAGED']),
        'user_role': user.role.name,
        'is_role_view': True,
    }
    
    return render(request, 'members/members/role_new_friends_list.html', context)

@login_required
def user_profile(request):
    """User profile page"""
    from members.forms import UserProfileForm
    from members.models import Attendance
    from django.db.models import Count
    from datetime import datetime, timedelta
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('members:user_profile')
    else:
        form = UserProfileForm(instance=request.user)
    
    # Calculate attendance summary for last 30 days
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_attendances = Attendance.objects.filter(
        user=request.user,
        date__gte=thirty_days_ago
    ).order_by('-date')
    
    total_attendances = recent_attendances.count()
    attendance_rate = (total_attendances / 30) * 100 if total_attendances > 0 else 0
    
    attendance_summary = {
        'total_attendances': total_attendances,
        'attendance_rate': round(attendance_rate, 1),
        'recent_attendances': recent_attendances[:10]
    }
    
    context = {
        'form': form,
        'attendance_summary': attendance_summary,
    }
    
    return render(request, 'members/auth/user_profile.html', context)

@login_required
def generate_qr_code(request, user_id):
    """Generate QR code for a user"""
    from django.core.exceptions import PermissionDenied
    
    user = get_object_or_404(CustomUser, id=user_id)

    if not (request.user.is_superuser or 
            request.user.role.name in ['ADMIN', 'VSL', 'CSL', 'CL'] or 
            request.user == user):
        raise PermissionDenied("You don't have permission to generate QR codes for this user.")
    
    try:
        qr_image = user.generate_qr_code()
        messages.success(request, f'QR code generated successfully for {user.full_name}')
    except Exception as e:
        messages.error(request, f'Error generating QR code: {str(e)}')
    
    if request.user == user:
        return redirect('members:user_profile')
    else:
        return redirect('members:member_detail', pk=user_id)

@login_required
def qr_scanner(request):
    """QR code scanner for attendance"""
    user = request.user
    church = user.church

    form = QRCodeScanForm()
    manual_form = ManualAttendanceForm(church=church)
    
    if request.method == 'POST':

        if 'manual_attendance' in request.POST:
            form = ManualAttendanceForm(request.POST, church=church)
            if form.is_valid():
                try:
                    attendee = form.cleaned_data['member']
                    date = form.cleaned_data['date']
                    time = form.cleaned_data['time']
                    service_type = form.cleaned_data['service_type']
                    notes = form.cleaned_data['notes']

                    existing_attendance = Attendance.objects.filter(
                        user=attendee,
                        date=date,
                        attendance_type=service_type
                    ).first()
                    
                    if existing_attendance:
                        messages.warning(request, f'{attendee.full_name} has already been marked present for {service_type} on {date}.')
                    else:
                        attendance = Attendance.objects.create(
                            user=attendee,
                            church=church,
                            attendance_type=service_type,
                            date=date,
                            time_in=time,
                            notes=notes,
                            scanned_by=user,
                            ip_address=request.META.get('REMOTE_ADDR'),
                            user_agent=request.META.get('HTTP_USER_AGENT', '')
                        )
                        
                        attendee.record_attendance()
                        
                        status_update_message = ""
                        if attendee.is_new_friend and service_type == 'SUNDAY':
                            current_status = attendee.timer_status
                            if current_status < 5:
                                new_status = current_status + 1
                                attendee.timer_status = new_status
                                attendee.save()
                                status_update_message = f" Status updated to {new_status}{'st' if new_status == 1 else 'nd' if new_status == 2 else 'rd' if new_status == 3 else 'th'} timer."
                            
                            elif current_status == 5:
                                from members.models import Role
                                cm_role, created = Role.objects.get_or_create(
                                    name='CM',
                                    defaults={'description': 'Care Member'}
                                )
                                attendee.transition_to_regular()
                                attendee.role = cm_role
                                attendee.save()
                                status_update_message = " Congratulations! You are now a regular member with CM role!"
                            else:
                                status_update_message = " You are already a regular member!"
                        
                        messages.success(request, f'Manual attendance recorded for {attendee.full_name}{status_update_message}')
                        
                        return JsonResponse({
                            'success': True,
                            'message': f'Manual attendance recorded for {attendee.full_name}{status_update_message}',
                            'user': {
                                'name': attendee.full_name,
                                'role': attendee.role.get_name_display() if attendee.role else 'No Role',
                                'time': attendance.time_in.strftime('%I:%M %p'),
                                'timer_status': attendee.timer_status if attendee.is_new_friend else None,
                                'is_new_friend': attendee.is_new_friend
                            }
                        })
                except Exception as e:
                    messages.error(request, f'Error recording manual attendance: {str(e)}')
                    return JsonResponse({
                        'success': False,
                        'message': f'Error recording manual attendance: {str(e)}'
                    })
        else:
            form = QRCodeScanForm(request.POST)
            if form.is_valid():
                qr_data = form.cleaned_data['qr_data']
                attendance_type = form.cleaned_data['attendance_type']
                notes = form.cleaned_data['notes']

                client_date_str = request.POST.get('client_date')
                client_time_str = request.POST.get('client_time')

                if client_date_str and client_time_str:
                    try:
                        from datetime import datetime
                        client_date = datetime.strptime(client_date_str, '%Y-%m-%d').date()
                        client_time = datetime.strptime(client_time_str, '%H:%M:%S').time()
                    except ValueError:
                        client_date = timezone.now().date()
                        client_time = timezone.now().time()
                else:
                    client_date = timezone.now().date()
                    client_time = timezone.now().time()
                
                try:
                    if qr_data.startswith('CHURCH_ATTENDANCE:'):
                        parts = qr_data.split(':')
                        if len(parts) >= 3:
                            qr_code_id = parts[1]
                            email = parts[2]

                            try:
                                attendee = CustomUser.objects.get(qr_code_id=qr_code_id, church=church)
                                existing_attendance = Attendance.objects.filter(
                                    user=attendee,
                                    date=client_date,
                                    attendance_type=attendance_type
                                ).first()
                                
                                if existing_attendance:
                                    messages.warning(request, f'{attendee.full_name} has already been marked present for {attendance_type} today.')
                                    return JsonResponse({
                                        'success': False,
                                        'message': 'qr is already scanned'
                                    })
                                else:
                                    attendance = Attendance.objects.create(
                                        user=attendee,
                                        church=church,
                                        attendance_type=attendance_type,
                                        date=client_date,     
                                        time_in=client_time, 
                                        notes=notes,
                                        scanned_by=user,
                                        ip_address=request.META.get('REMOTE_ADDR'),
                                        user_agent=request.META.get('HTTP_USER_AGENT', '')
                                    )

                                    attendee.record_attendance()

                                    status_update_message = ""
                                    if attendee.is_new_friend and attendance_type == 'SUNDAY':
                                        current_status = attendee.timer_status
                                        if current_status < 5:
                                            new_status = current_status + 1
                                            attendee.timer_status = new_status
                                            attendee.save()
                                            status_update_message = f" Status updated to {new_status}{'st' if new_status == 1 else 'nd' if new_status == 2 else 'rd' if new_status == 3 else 'th'} timer."
                                        
                                        elif current_status == 5:
                                            from members.models import Role
                                            cm_role, created = Role.objects.get_or_create(
                                                name='CM',
                                                defaults={'description': 'Care Member'}
                                            )
                                            attendee.transition_to_regular()
                                            attendee.role = cm_role
                                            attendee.save()
                                            status_update_message = " Congratulations! You are now a regular member with CM role!"
                                        else:
                                            status_update_message = " You are already a regular member!"
                                    
                                    messages.success(request, f'Attendance recorded for {attendee.full_name}{status_update_message}')
                                    
                                    return JsonResponse({
                                        'success': True,
                                        'message': f'Attendance recorded for {attendee.full_name}{status_update_message}',
                                        'user': {
                                            'name': attendee.full_name,
                                            'role': attendee.role.get_name_display() if attendee.role else 'No Role',
                                            'time': attendance.time_in.strftime('%I:%M %p'),
                                            'timer_status': attendee.timer_status if attendee.is_new_friend else None,
                                            'is_new_friend': attendee.is_new_friend
                                        }
                                    })
                            except CustomUser.DoesNotExist:
                                messages.error(request, 'User not found or QR code is invalid.')
                                return JsonResponse({
                                    'success': False,
                                    'message': 'User not found or QR code is invalid.'
                                })
                        else:
                            messages.error(request, 'Invalid QR code format.')
                            return JsonResponse({
                                'success': False,
                                'message': 'Invalid QR code format.'
                            })
                    else:
                        messages.error(request, 'Invalid QR code. Please scan a valid church attendance QR code.')
                        return JsonResponse({
                            'success': False,
                            'message': 'Invalid QR code. Please scan a valid church attendance QR code.'
                        })
                except Exception as e:
                    messages.error(request, f'Error processing QR code: {str(e)}')
                    return JsonResponse({
                        'success': False,
                        'message': f'Error processing QR code: {str(e)}'
                    })

    recent_attendances = Attendance.objects.filter(
        church=church,
        date=timezone.now().date()
    ).select_related('user').order_by('-time_in')[:10]
    
    context = {
        'form': form,
        'manual_form': manual_form,
        'recent_attendances': recent_attendances,
    }
    
    return render(request, 'members/attendance/qr_scanner.html', context)

@login_required
def attendance_list(request):
    """List attendance records with filtering"""
    user = request.user
    church = user.church

    form = AttendanceFilterForm(request.GET, church=church)
    attendances = Attendance.objects.filter(church=church).select_related('user', 'scanned_by')
    
    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        attendance_type = form.cleaned_data.get('attendance_type')
        user_filter = form.cleaned_data.get('user')
        
        if date_from:
            attendances = attendances.filter(date__gte=date_from)
        if date_to:
            attendances = attendances.filter(date__lte=date_to)
        if attendance_type:
            attendances = attendances.filter(attendance_type=attendance_type)
        if user_filter:
            attendances = attendances.filter(user=user_filter)

    paginator = Paginator(attendances, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    today = timezone.now().date()
    today_attendances = Attendance.objects.filter(church=church, date=today).count()
    this_week = Attendance.objects.filter(
        church=church, 
        date__gte=today - timedelta(days=7)
    ).count()
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'today_attendances': today_attendances,
        'this_week_attendances': this_week,
    }
    
    return render(request, 'members/attendance/attendance_list.html', context)

@login_required
def attendance_export(request):
    """Export attendance data"""
    user = request.user
    church = user.church
    
    if request.method == 'POST':
        form = AttendanceExportForm(request.POST)
        if form.is_valid():
            export_format = form.cleaned_data['format']
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
            attendance_type = form.cleaned_data.get('attendance_type')
            include_qr_codes = form.cleaned_data.get('include_qr_codes', False)

            attendances = Attendance.objects.filter(church=church).select_related('user', 'scanned_by')
            
            if date_from:
                attendances = attendances.filter(date__gte=date_from)
            if date_to:
                attendances = attendances.filter(date__lte=date_to)
            if attendance_type:
                attendances = attendances.filter(attendance_type=attendance_type)

            if export_format == 'csv':
                from django.http import HttpResponse
                import csv
                
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="attendance_{church.domain}_{timezone.now().strftime("%Y%m%d")}.csv"'
                
                writer = csv.writer(response)
                headers = ['Date', 'Time In', 'Time Out', 'Name', 'Email', 'Role', 'Attendance Type', 'Notes', 'Scanned By']
                if include_qr_codes:
                    headers.append('QR Code ID')
                writer.writerow(headers)
                
                for attendance in attendances:
                    row = [
                        attendance.date.strftime('%Y-%m-%d'),
                        attendance.time_in.strftime('%H:%M:%S') if attendance.time_in else '',
                        attendance.time_out.strftime('%H:%M:%S') if attendance.time_out else '',
                        attendance.user.full_name,
                        attendance.user.email,
                        attendance.user.role.get_name_display() if attendance.user.role else '',
                        attendance.get_attendance_type_display(),
                        attendance.notes or '',
                        attendance.scanned_by.full_name if attendance.scanned_by else ''
                    ]
                    if include_qr_codes:
                        row.append(attendance.user.qr_code_id or '')
                    writer.writerow(row)
                
                return response
    else:
        form = AttendanceExportForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'members/attendance/attendance_export.html', context)

@login_required
def profile_export(request):
    """Export user profiles with QR codes"""
    user = request.user
    church = user.church
    
    if request.method == 'POST':
        form = ProfileExportForm(request.POST)
        if form.is_valid():
            export_format = form.cleaned_data['format']
            include_qr_codes = form.cleaned_data.get('include_qr_codes', False)
            include_profile_pictures = form.cleaned_data.get('include_profile_pictures', False)
            member_type = form.cleaned_data.get('member_type', '')
            
            users = CustomUser.objects.filter(church=church, is_active=True)
            
            if member_type == 'new_friends':
                users = users.filter(is_new_friend=True)
            elif member_type == 'regular_members':
                users = users.filter(is_new_friend=False)

            if export_format == 'csv':
                from django.http import HttpResponse
                import csv
                
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="profiles_{church.domain}_{timezone.now().strftime("%Y%m%d")}.csv"'
                
                writer = csv.writer(response)
                headers = ['Name', 'Email', 'Phone', 'Address', 'Birth Date', 'Role', 'Member Type', 'Date Joined']
                if include_qr_codes:
                    headers.append('QR Code ID')
                writer.writerow(headers)
                
                for user_obj in users:
                    row = [
                        user_obj.full_name,
                        user_obj.email,
                        user_obj.phone_number,
                        user_obj.address,
                        user_obj.birth_date.strftime('%Y-%m-%d') if user_obj.birth_date else '',
                        user_obj.role.get_name_display() if user_obj.role else '',
                        'New Friend' if user_obj.is_new_friend else 'Regular Member',
                        user_obj.date_joined.strftime('%Y-%m-%d')
                    ]
                    if include_qr_codes:
                        row.append(str(user_obj.qr_code_id))
                    writer.writerow(row)
                
                return response
            
            elif export_format == 'excel':
                import pandas as pd
                from io import BytesIO
                
                # Prepare data for export
                data = []
                for user_obj in users:
                    row_data = {
                        'Name': user_obj.full_name,
                        'Email': user_obj.email,
                        'Phone': user_obj.phone_number,
                        'Address': user_obj.address,
                        'Birth Date': user_obj.birth_date.strftime('%Y-%m-%d') if user_obj.birth_date else '',
                        'Role': user_obj.role.get_name_display() if user_obj.role else '',
                        'Member Type': 'New Friend' if user_obj.is_new_friend else 'Regular Member',
                        'Date Joined': user_obj.date_joined.strftime('%Y-%m-%d')
                    }
                    if include_qr_codes:
                        row_data['QR Code ID'] = str(user_obj.qr_code_id)
                    data.append(row_data)
                
                # Create DataFrame from data
                df = pd.DataFrame(data)
                
                # Create in-memory output file for Excel
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Profiles')
                    
                    # Get the workbook and worksheet to apply formatting
                    workbook = writer.book
                    worksheet = writer.sheets['Profiles']
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Set response content
                output.seek(0)
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="profiles_{church.domain}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
                return response
            
            elif export_format == 'pdf':
                messages.info(request, 'PDF export feature will be implemented soon.')
                return redirect('members:profile_export')
    else:
        form = ProfileExportForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'members/reports/profile_export.html', context)

@login_required
def profile_import(request):
    """Import user profiles"""
    user = request.user
    church = user.church
    
    if request.method == 'POST':
        form = ProfileImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            update_existing = form.cleaned_data.get('update_existing', False)
            generate_qr_codes = form.cleaned_data.get('generate_qr_codes', True)
            
            try:
                if file.name.endswith('.csv'):
                    import csv
                    import io
                    from datetime import datetime
                    
                    decoded_file = file.read().decode('utf-8')
                    csv_data = csv.DictReader(io.StringIO(decoded_file))
                    
                    imported_count = 0
                    updated_count = 0
                    
                    for row in csv_data:
                        email = row.get('email', '').strip()
                        if not email:
                            continue
                        try:
                            existing_user = CustomUser.objects.get(email=email, church=church)
                            
                            if update_existing:
                                existing_user.first_name = row.get('first_name', existing_user.first_name)
                                existing_user.last_name = row.get('last_name', existing_user.last_name)
                                existing_user.phone_number = row.get('phone_number', existing_user.phone_number)
                                existing_user.address = row.get('address', existing_user.address)
                                if row.get('birth_date'):
                                    try:
                                        existing_user.birth_date = datetime.strptime(row['birth_date'], '%Y-%m-%d').date()
                                    except ValueError:
                                        pass
                                existing_user.save()
                                updated_count += 1
                        except CustomUser.DoesNotExist:
                            new_user = CustomUser.objects.create(
                                email=email,
                                first_name=row.get('first_name', ''),
                                last_name=row.get('last_name', ''),
                                phone_number=row.get('phone_number', ''),
                                address=row.get('address', ''),
                                church=church,
                                is_new_friend=True,
                                is_active=True
                            )
                            
                            if row.get('birth_date'):
                                try:
                                    new_user.birth_date = datetime.strptime(row['birth_date'], '%Y-%m-%d').date()
                                    new_user.save()
                                except ValueError:
                                    pass

                            if generate_qr_codes:
                                new_user.generate_qr_code()
                            
                            imported_count += 1
                    
                    messages.success(request, f'Import completed: {imported_count} new users imported, {updated_count} users updated.')
                else:
                    messages.info(request, 'Excel import feature will be implemented soon.')
                
            except Exception as e:
                messages.error(request, f'Error importing file: {str(e)}')
    else:
        form = ProfileImportForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'members/reports/profile_import.html', context)

@login_required
def care_group_report_list(request):
    """List all care group reports for the user's church"""
    user = request.user
    church = user.church

    if not user.role or user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to view care group reports.')
        return redirect('churches:dashboard')

    search = request.GET.get('search', '')
    care_group_filter = request.GET.get('care_group', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    reports = CareGroupReport.objects.filter(church=church)

    if user.role.name == 'VSL':
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        reports = reports.filter(
            Q(care_group__leader=user) | 
            Q(care_group__leader__in=member_users)
        ).distinct()
    elif user.role.name == 'CSL':
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        reports = reports.filter(
            Q(care_group__leader=user) | 
            Q(care_group__leader__in=member_users)
        ).distinct()
    elif user.role.name == 'CL':
        reports = reports.filter(care_group__leader=user)
        
    elif user.role.name == 'ADMIN':
        pass

    if search:
        reports = reports.filter(
            Q(care_group__name__icontains=search) |
            Q(topic_discussed__icontains=search) |
            Q(scripture_used__icontains=search)
        )
    
    if care_group_filter:
        reports = reports.filter(care_group_id=care_group_filter)
    
    if date_from:
        reports = reports.filter(date_of_cg__gte=date_from)
    
    if date_to:
        reports = reports.filter(date_of_cg__lte=date_to)

    reports = reports.order_by('-date_of_cg', '-created_at')

    paginator = Paginator(reports, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    care_groups = Group.objects.filter(
        church=church,
        group_type='CARE',
        is_active=True
    )

    if user.role.name == 'VSL':
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        care_groups = care_groups.filter(
            Q(leader=user) | 
            Q(leader__in=member_users)
        ).distinct()
        
    elif user.role.name == 'CSL':
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        care_groups = care_groups.filter(
            Q(leader=user) | 
            Q(leader__in=member_users)  
        ).distinct()
    elif user.role.name == 'CL':
        care_groups = care_groups.filter(leader=user)
    elif user.role.name == 'ADMIN':
        pass
    
    care_groups = care_groups.order_by('name')
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'care_group_filter': care_group_filter,
        'date_from': date_from,
        'date_to': date_to,
        'care_groups': care_groups,
        'total_reports': reports.count(),
    }
    
    return render(request, 'members/groups/care_group_report_list.html', context)

@login_required
def care_group_report_create(request):
    """Create care group report"""
    user = request.user

    if user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to create care group reports.')
        return redirect('members:care_group_report_list')
    
    if request.method == 'POST':
        form = CareGroupReportForm(request.POST, user=user)
        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = user
            report.church = user.church
            report.save()

            ActivityLog.objects.create(
                user=user,
                church=user.church,
                action='REPORT_CREATED',
                description=f'Created care group report: {report.care_group.name} - {report.date_of_cg}'
            )
            
            messages.success(request, f'Care group report created successfully!')
            return redirect('members:care_group_member_report_create', report_id=report.pk)
    else:
        form = CareGroupReportForm(user=user)
    
    context = {
        'form': form,
        'title': 'Create Care Group Report',
        'user_role': user.role.name,
    }
    
    return render(request, 'members/groups/care_group_report_form.html', context)

@login_required
def care_group_member_report_create(request, report_id):
    """Create care group member report"""
    user = request.user
    church = user.church

    if not user.role or user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to create care group member reports.')
        return redirect('churches:dashboard')

    try:
        report = CareGroupReport.objects.get(id=report_id, church=church)
    except CareGroupReport.DoesNotExist:
        messages.error(request, 'Care group report not found.')
        return redirect('members:care_group_report_list')

    # Check permissions
    if user.role.name == 'CL' and report.care_group.leader != user:
        messages.error(request, 'You can only create member reports for your own care groups.')
        return redirect('members:care_group_report_list')
    elif user.role.name in ['VSL', 'CSL']:
        # Check if user can access this report based on their leadership hierarchy
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        if not (report.care_group.leader == user or report.care_group.leader in member_users):
            messages.error(request, 'You can only create member reports for care groups under your leadership.')
            return redirect('members:care_group_report_list')

    # Get care group members
    care_group_members = CustomUser.objects.filter(
        regular_member_profile__group=report.care_group,
        is_active=True
    ).order_by('first_name', 'last_name')

    # Get existing member reports for this care group report
    existing_reports = {}
    for member_report in report.member_reports.all():
        existing_reports[member_report.member.id] = member_report

    if request.method == 'POST':
        # Process form submission
        for member in care_group_members:
            member_id = str(member.id)
            
            # Get or create member report
            member_report, created = CareGroupMemberReport.objects.get_or_create(
                report=report,
                member=member,
                defaults={
                    'status': 'ACTIVE',
                    'sunday_attendance': False,
                    'group_attendance': False,
                    'new_disciples_invited': 0,
                    'follow_ups': 0,
                    'notes': ''
                }
            )
            
            # Update fields based on form data
            member_report.sunday_attendance = f'sunday_attendance_{member_id}' in request.POST
            member_report.group_attendance = f'group_attendance_{member_id}' in request.POST
            member_report.new_disciples_invited = int(request.POST.get(f'new_disciples_invited_{member_id}', 0))
            member_report.follow_ups = int(request.POST.get(f'follow_ups_{member_id}', 0))
            member_report.notes = request.POST.get(f'notes_{member_id}', '')
            member_report.save()

        messages.success(request, 'Member reports saved successfully!')
        return redirect('members:care_group_report_detail', report_id=report.id)
    
    context = {
        'report': report,
        'care_group_members': care_group_members,
        'existing_reports': existing_reports,
        'user_role': user.role.name,
    }
    
    return render(request, 'members/groups/care_group_member_report_form.html', context)

@login_required
def care_group_report_detail(request, report_id):
    """Care group report detail"""
    user = request.user
    church = user.church

    if not user.role or user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to view care group reports.')
        return redirect('churches:dashboard')

    try:
        report = CareGroupReport.objects.get(id=report_id, church=church)
    except CareGroupReport.DoesNotExist:
        messages.error(request, 'Care group report not found.')
        return redirect('members:care_group_report_list')

    # Check permissions
    if user.role.name == 'CL' and report.care_group.leader != user:
        messages.error(request, 'You can only view reports for your own care groups.')
        return redirect('members:care_group_report_list')
    elif user.role.name in ['VSL', 'CSL']:
        # Check if user can access this report based on their leadership hierarchy
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        if not (report.care_group.leader == user or report.care_group.leader in member_users):
            messages.error(request, 'You can only view reports for care groups under your leadership.')
            return redirect('members:care_group_report_list')

    member_reports = report.member_reports.select_related('member').order_by('member__first_name', 'member__last_name')
    
    context = {
        'report': report,
        'member_reports': member_reports,
        'user_role': user.role.name,
    }
    
    return render(request, 'members/groups/care_group_report_detail.html', context)

@login_required
def care_group_report_print(request, report_id):
    """Print care group report"""
    user = request.user
    church = user.church

    if not user.role or user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to print care group reports.')
        return redirect('churches:dashboard')

    try:
        report = CareGroupReport.objects.get(id=report_id, church=church)
    except CareGroupReport.DoesNotExist:
        messages.error(request, 'Care group report not found.')
        return redirect('members:care_group_report_list')

    # Check permissions
    if user.role.name == 'CL' and report.care_group.leader != user:
        messages.error(request, 'You can only print reports for your own care groups.')
        return redirect('members:care_group_report_list')
    elif user.role.name in ['VSL', 'CSL']:
        # Check if user can access this report based on their leadership hierarchy
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        if not (report.care_group.leader == user or report.care_group.leader in member_users):
            messages.error(request, 'You can only print reports for care groups under your leadership.')
            return redirect('members:care_group_report_list')

    member_reports = report.member_reports.select_related('member').order_by('member__first_name', 'member__last_name')
    
    context = {
        'report': report,
        'member_reports': member_reports,
        'user_role': user.role.name,
    }
    
    return render(request, 'members/groups/care_group_report_print.html', context)

@login_required
def care_group_report_edit(request, report_id):
    """Edit care group report"""
    user = request.user
    church = user.church

    if not user.role or user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to edit care group reports.')
        return redirect('churches:dashboard')

    try:
        report = CareGroupReport.objects.get(id=report_id, church=church)
    except CareGroupReport.DoesNotExist:
        messages.error(request, 'Care group report not found.')
        return redirect('members:care_group_report_list')

    # Check permissions - admin can edit all reports, others can only edit their own
    if user.role.name != 'ADMIN' and report.created_by != user:
        messages.error(request, 'You can only edit reports you created.')
        return redirect('members:care_group_report_detail', report_id=report_id)

    if request.method == 'POST':
        # Handle form submission
        form = CareGroupReportForm(request.POST, instance=report, user=user)
        if form.is_valid():
            updated_report = form.save()
            
            ActivityLog.objects.create(
                user=user,
                church=user.church,
                action='REPORT_UPDATED',
                description=f'Updated care group report: {updated_report.care_group.name} - {updated_report.date_of_cg}'
            )
            
            messages.success(request, f'Care group report updated successfully!')
            return redirect('members:care_group_report_detail', report_id=report_id)
    else:
        form = CareGroupReportForm(instance=report, user=user)
    
    context = {
        'form': form,
        'report': report,
        'user_role': user.role.name,
    }
    
    return render(request, 'members/groups/care_group_report_edit.html', context)

@login_required
def care_group_report_delete(request, report_id):
    """Delete care group report"""
    user = request.user
    church = user.church

    if not user.role or user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to delete care group reports.')
        return redirect('churches:dashboard')

    try:
        report = CareGroupReport.objects.get(id=report_id, church=church)
    except CareGroupReport.DoesNotExist:
        messages.error(request, 'Care group report not found.')
        return redirect('members:care_group_report_list')

    # Check permissions - admin can delete all reports, others can only delete their own
    if user.role.name != 'ADMIN' and report.created_by != user:
        messages.error(request, 'You can only delete reports you created.')
        return redirect('members:care_group_report_detail', report_id=report_id)

    if request.method == 'POST':
        report_name = report.care_group.name
        report_date = report.date_of_cg
        report.delete()
        
        ActivityLog.objects.create(
            user=user,
            church=user.church,
            action='REPORT_DELETED',
            description=f'Deleted care group report: {report_name} - {report_date}'
        )
        
        messages.success(request, f'Care group report deleted successfully!')
        return redirect('members:care_group_report_list')
    
    context = {
        'report': report,
    }
    
    return render(request, 'members/groups/care_group_report_confirm_delete.html', context)

@login_required
def care_group_attendance_tracking(request, group_id):
    """Care group attendance tracking"""
    return render(request, 'members/attendance/care_group_attendance_tracking.html', {})

@login_required
def care_group_report_export(request):
    """Export care group reports to CSV or Excel"""
    user = request.user
    church = user.church

    if not user.role or user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to export care group reports.')
        return redirect('churches:dashboard')

    format_type = request.GET.get('format', 'csv')
    care_group_id = request.GET.get('care_group', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    reports = CareGroupReport.objects.filter(church=church).select_related(
        'care_group', 'vine_servant_leader', 'cluster_servant_leader', 'care_leader', 'created_by'
    ).prefetch_related('member_reports__member')

    if user.role.name == 'VSL':
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        reports = reports.filter(
            Q(care_group__leader=user) | 
            Q(care_group__leader__in=member_users)
        ).distinct()
    elif user.role.name == 'CSL':
        user_groups = Group.objects.filter(leader=user, group_type='CARE')
        member_users = CustomUser.objects.filter(
            regular_member_profile__group__in=user_groups
        ).exclude(id=user.id)
        
        reports = reports.filter(
            Q(care_group__leader=user) | 
            Q(care_group__leader__in=member_users)
        ).distinct()
    elif user.role.name == 'CL':
        reports = reports.filter(care_group__leader=user)

    if care_group_id:
        reports = reports.filter(care_group_id=care_group_id)
    
    if date_from:
        reports = reports.filter(date_of_cg__gte=date_from)
    
    if date_to:
        reports = reports.filter(date_of_cg__lte=date_to)

    reports = reports.order_by('-date_of_cg', '-created_at')

    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="care_group_reports_{church.name}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Report ID', 'Care Group', 'Date of CG', 'Topic Discussed', 'Scripture Used',
            'Group Day', 'Group Time', 'Venue Address', 'Contact Number',
            'Vine Servant Leader', 'Cluster Servant Leader', 'Care Leader',
            'Total Members Reported', 'Total Sunday Attendance', 'Total Group Attendance',
            'Total New Disciples Invited', 'Total Follow-ups', 'Created By', 'Created At'
        ])
        
        # Write data
        for report in reports:
            writer.writerow([
                report.id,
                report.care_group.name,
                report.date_of_cg,
                report.topic_discussed,
                report.scripture_used,
                report.get_group_day_display() if report.group_day else '',
                report.group_time.strftime('%H:%M') if report.group_time else '',
                report.venue_address,
                report.contact_number,
                report.vine_servant_leader.full_name if report.vine_servant_leader else '',
                report.cluster_servant_leader.full_name if report.cluster_servant_leader else '',
                report.care_leader.full_name if report.care_leader else '',
                report.total_members_reported,
                report.total_sunday_attendance,
                report.total_group_attendance,
                report.total_new_disciples_invited,
                report.total_follow_ups,
                report.created_by.full_name,
                report.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    elif format_type == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl package. Please install it.')
            return redirect('members:care_group_report_list')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Care Group Reports"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header
        headers = [
            'Report ID', 'Care Group', 'Date of CG', 'Topic Discussed', 'Scripture Used',
            'Group Day', 'Group Time', 'Venue Address', 'Contact Number',
            'Vine Servant Leader', 'Cluster Servant Leader', 'Care Leader',
            'Total Members Reported', 'Total Sunday Attendance', 'Total Group Attendance',
            'Total New Disciples Invited', 'Total Follow-ups', 'Created By', 'Created At'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, report in enumerate(reports, 2):
            ws.cell(row=row, column=1, value=report.id)
            ws.cell(row=row, column=2, value=report.care_group.name)
            ws.cell(row=row, column=3, value=report.date_of_cg)
            ws.cell(row=row, column=4, value=report.topic_discussed)
            ws.cell(row=row, column=5, value=report.scripture_used)
            ws.cell(row=row, column=6, value=report.get_group_day_display() if report.group_day else '')
            ws.cell(row=row, column=7, value=report.group_time.strftime('%H:%M') if report.group_time else '')
            ws.cell(row=row, column=8, value=report.venue_address)
            ws.cell(row=row, column=9, value=report.contact_number)
            ws.cell(row=row, column=10, value=report.vine_servant_leader.full_name if report.vine_servant_leader else '')
            ws.cell(row=row, column=11, value=report.cluster_servant_leader.full_name if report.cluster_servant_leader else '')
            ws.cell(row=row, column=12, value=report.care_leader.full_name if report.care_leader else '')
            ws.cell(row=row, column=13, value=report.total_members_reported)
            ws.cell(row=row, column=14, value=report.total_sunday_attendance)
            ws.cell(row=row, column=15, value=report.total_group_attendance)
            ws.cell(row=row, column=16, value=report.total_new_disciples_invited)
            ws.cell(row=row, column=17, value=report.total_follow_ups)
            ws.cell(row=row, column=18, value=report.created_by.full_name)
            ws.cell(row=row, column=19, value=report.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="care_group_reports_{church.name}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    else:
        messages.error(request, 'Invalid format specified.')
        return redirect('members:care_group_report_list')

@login_required
def care_group_member_report_export(request, report_id):
    """Export individual care group member reports to CSV or Excel"""
    user = request.user
    church = user.church

    if not user.role or user.role.name not in ['VSL', 'CSL', 'CL', 'ADMIN']:
        messages.error(request, 'You do not have permission to export care group member reports.')
        return redirect('churches:dashboard')

    try:
        report = CareGroupReport.objects.get(id=report_id, church=church)
    except CareGroupReport.DoesNotExist:
        messages.error(request, 'Care group report not found.')
        return redirect('members:care_group_report_list')

    # Check permissions
    if user.role.name == 'CL' and report.care_group.leader != user:
        messages.error(request, 'You can only export reports for your own care groups.')
        return redirect('members:care_group_report_list')

    format_type = request.GET.get('format', 'csv')
    member_reports = report.member_reports.select_related('member').order_by('member__first_name', 'member__last_name')

    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="care_group_member_reports_{report.care_group.name}_{report.date_of_cg}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Member Name', 'Status', 'Sunday Attendance', 'Group Attendance',
            'New Disciples Invited', 'Follow-ups', 'Notes'
        ])
        
        # Write data
        for member_report in member_reports:
            writer.writerow([
                member_report.member.full_name,
                member_report.get_status_display(),
                'Yes' if member_report.sunday_attendance else 'No',
                'Yes' if member_report.group_attendance else 'No',
                member_report.new_disciples_invited,
                member_report.follow_ups,
                member_report.notes
            ])
        
        return response
    
    elif format_type == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl package. Please install it.')
            return redirect('members:care_group_report_detail', report_id=report_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Member Reports - {report.care_group.name}"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header
        headers = [
            'Member Name', 'Status', 'Sunday Attendance', 'Group Attendance',
            'New Disciples Invited', 'Follow-ups', 'Notes'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, member_report in enumerate(member_reports, 2):
            ws.cell(row=row, column=1, value=member_report.member.full_name)
            ws.cell(row=row, column=2, value=member_report.get_status_display())
            ws.cell(row=row, column=3, value='Yes' if member_report.sunday_attendance else 'No')
            ws.cell(row=row, column=4, value='Yes' if member_report.group_attendance else 'No')
            ws.cell(row=row, column=5, value=member_report.new_disciples_invited)
            ws.cell(row=row, column=6, value=member_report.follow_ups)
            ws.cell(row=row, column=7, value=member_report.notes)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="care_group_member_reports_{report.care_group.name}_{report.date_of_cg}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    else:
        messages.error(request, 'Invalid format specified.')
        return redirect('members:care_group_report_detail', report_id=report_id)
